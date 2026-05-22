"""
processor.py
────────────────────────────────────────────────────────────────────────────
Core logic for building the Image Stack Import Template.

Row-level failure logic
───────────────────────
If ANY URL in a source row fails to resolve (returns a UUID-like name with no
extension), the ENTIRE row is treated as failed. This preserves stack order
integrity — a partial row would renumber the remaining images incorrectly.

Two-phase architecture
──────────────────────
Phase 1  fetch_all_filenames()  →  resolves every URL, marks failures
Phase 2  write_output_sheet()   →  writes output with skip_failed choice
         highlight_source_failures()  →  paints red cells in Image Links sheet

The caller (app.py) runs Phase 1 once and caches results, then runs Phase 2
when the user confirms their choice — no re-fetching required.
"""

import re
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────── constants ───────────────────────────────────────

IMPORT_TYPE = "Create/Edit"
HEADERS_OUT = ["Import Type", "Collection Folder",
               "Image Stack Group", "Filename", "Image Stack Order"]

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept":          "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control":   "no-cache",
    "Pragma":          "no-cache",
}

MIME_EXT = {
    "image/jpeg": "jpg", "image/jpg": "jpg", "image/png": "png",
    "image/webp": "webp", "image/avif": "avif", "image/gif": "gif",
    "image/svg+xml": "svg", "image/bmp": "bmp",
    "image/tiff": "tif", "image/tif": "tif",
}

BAD_CHARS  = re.compile(r'[\\/:*?"<>|]')
UUID_LIKE  = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.IGNORECASE)

RED_FILL      = PatternFill("solid", fgColor="922B21")   # dark crimson — only on failed cells
RED_FONT      = Font(color="FDFEFE", bold=True)           # near-white text on dark bg
HEADER_FILL   = PatternFill("solid", fgColor="1F2D4E")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


# ─────────────────────────── helpers ─────────────────────────────────────────

def sanitize(name: str) -> str:
    name = BAD_CHARS.sub("_", name)
    return name.rstrip(" .") or "download"

def url_origin(url: str) -> str:
    p = urllib.parse.urlparse(url)
    return f"{p.scheme}://{p.netloc}/"

def filename_from_url(url: str) -> str:
    path = urllib.parse.urlparse(url).path
    name = urllib.parse.unquote(path.rstrip("/").rsplit("/", 1)[-1])
    return name or "download"

def filename_from_content_disposition(cd: str) -> str:
    if not cd:
        return ""
    m = re.search(r"filename\*\s*=\s*([^;]+)", cd, re.IGNORECASE)
    if m:
        val = m.group(1).strip().strip("'\"")
        if "''" in val:
            val = val.split("''", 1)[1]
        return urllib.parse.unquote(val)
    m = re.search(r'filename\s*=\s*"?([^";]+)"?', cd, re.IGNORECASE)
    return m.group(1).strip().strip('"') if m else ""

def guess_ext(content_type: str) -> str:
    return MIME_EXT.get((content_type or "").split(";")[0].strip().lower(), "")

def _looks_failed(name: str) -> bool:
    """True when the name has no file extension AND looks like a bare UUID slug."""
    if "." in name:
        return False
    if name in ("download",):
        return True
    if UUID_LIKE.match(name):
        return True
    return False

def _parse_headers(resp, original_url: str) -> tuple[str, bool]:
    http_ok = resp.status_code < 400
    cd  = resp.headers.get("Content-Disposition", "")
    ct  = resp.headers.get("Content-Type", "")
    name = filename_from_content_disposition(cd) or filename_from_url(resp.url or original_url)
    name = sanitize(urllib.parse.unquote(name))
    if "." not in name:
        ext = guess_ext(ct)
        if ext:
            name = f"{name}.{ext}"
    return (name or "download"), http_ok

def get_filename(url: str, referer: str | None, timeout: int,
                 session: requests.Session) -> dict:
    headers = dict(BROWSER_HEADERS)
    headers["Referer"] = referer or url_origin(url)
    try:
        resp = session.head(url, headers=headers, timeout=timeout, allow_redirects=True)
        if resp.status_code not in (405, 403):
            name, http_ok = _parse_headers(resp, url)
            if name and name != "download" and "." in name:
                return {"filename": name, "failed": not http_ok, "url": url}
        resp = session.get(url, headers=headers, timeout=timeout, allow_redirects=True, stream=True)
        resp.close()
        name, http_ok = _parse_headers(resp, url)
        failed = (not http_ok) or _looks_failed(name)
        return {"filename": name, "failed": failed, "url": url}
    except Exception:
        fallback = sanitize(filename_from_url(url)) or "download"
        return {"filename": fallback, "failed": _looks_failed(fallback), "url": url}


# ─────────────────────────── core logic ──────────────────────────────────────

def collect_work(ws_src, first_data_row: int = 2):
    """
    Returns:
        rows_by_src_row  – dict{ src_row: [(master_id, stack_group, stack_order, url, src_col), ...] }
        urls             – deduplicated set of all URLs
    """
    max_row = ws_src.max_row
    max_col = ws_src.max_column
    rows_by_src_row = {}
    urls = set()

    for r in range(first_data_row, max_row + 1):
        master_id = ws_src.cell(r, 1).value or ""
        if isinstance(master_id, float) and master_id.is_integer():
            master_id = str(int(master_id))
        master_id   = str(master_id).strip()
        stack_group = str(ws_src.cell(r, 2).value or "").strip()
        if not master_id or not stack_group:
            continue

        cells = []
        for c in range(3, max_col + 1):
            url = str(ws_src.cell(r, c).value or "").strip()
            if url:
                cells.append((master_id, stack_group, c - 2, url, c))
                urls.add(url)

        if cells:
            rows_by_src_row[r] = cells

    return rows_by_src_row, urls


def fetch_all_filenames(urls: set, referer, timeout, workers, progress_callback=None) -> dict:
    cache = {}
    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(pool_connections=workers, pool_maxsize=workers, max_retries=1)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    total, done, t0 = len(urls), 0, time.time()

    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {pool.submit(get_filename, url, referer, timeout, session): url for url in urls}
        for future in as_completed(future_map):
            url = future_map[future]
            cache[url] = future.result()
            done += 1
            if progress_callback:
                progress_callback(done, total, time.time() - t0)

    session.close()
    return cache


def build_row_results(rows_by_src_row: dict, cache: dict):
    """
    Returns:
        good_rows   – list of (master_id, stack_group, stack_order, url, src_row, src_col)
                      only rows where EVERY url succeeded
        failed_rows – list of { master_id, stack_group, urls: [{ url, filename }], src_row }
                      rows where at least ONE url failed (entire row is failed)
        failed_src_cells – set of (src_row, src_col) to highlight red in source sheet
    """
    good_rows        = []
    failed_rows      = []
    failed_src_cells = set()

    for src_row, cells in sorted(rows_by_src_row.items()):
        # Check if ANY url in this row failed
        row_has_failure = any(cache.get(url, {}).get("failed", False) for (_mid, _sg, _so, url, _sc) in cells)

        if row_has_failure:
            # Collect all failed URLs in this row (for display)
            failed_url_info = []
            for (master_id, stack_group, stack_order, url, src_col) in cells:
                entry = cache.get(url, {})
                failed_url_info.append({
                    "url":         url,
                    "filename":    entry.get("filename", ""),
                    "failed":      entry.get("failed", False),
                    "stack_order": stack_order,
                })
                # Only highlight the specific cells whose URL actually failed
                if entry.get("failed", False):
                    failed_src_cells.add((src_row, src_col))

            # Use first cell's master_id / stack_group for display
            first = cells[0]
            failed_rows.append({
                "master_id":   first[0],
                "stack_group": first[1],
                "src_row":     src_row,
                "urls":        failed_url_info,
            })
        else:
            for (master_id, stack_group, stack_order, url, src_col) in cells:
                good_rows.append((master_id, stack_group, stack_order, url, src_row, src_col))

    return good_rows, failed_rows, failed_src_cells


def highlight_source_failures(ws_src, failed_src_cells: set):
    """Paint every URL cell in a failed row red in the source sheet."""
    for (src_row, src_col) in failed_src_cells:
        cell = ws_src.cell(src_row, src_col)
        cell.fill = RED_FILL
        cell.font = RED_FONT


def write_output_sheet(wb, sheet_name: str, rows_to_write: list, cache: dict):
    """Write only good rows to the output sheet."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    for col_idx, hdr in enumerate(HEADERS_OUT, start=1):
        cell = ws.cell(1, col_idx, hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for out_row, (master_id, stack_group, stack_order, url, _sr, _sc) in enumerate(rows_to_write, start=2):
        fname = cache.get(url, {}).get("filename", "download")
        ws.cell(out_row, 1, IMPORT_TYPE)
        ws.cell(out_row, 2, master_id)
        ws.cell(out_row, 3, stack_group)
        ws.cell(out_row, 4, fname)
        ws.cell(out_row, 5, stack_order)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    return ws


def process_workbook(file_path, output_path,
                     sheet_src="Image Links",
                     sheet_out="Image Stack Import Template",
                     workers=25, timeout=10, referer=None,
                     skip_failed=True,
                     progress_callback=None,
                     cached_data=None):
    """
    Full pipeline.

    cached_data: if provided (dict with keys 'cache', 'rows_by_src_row'),
                 skips the HTTP fetch phase entirely. Used by /confirm to
                 re-write the output with a different skip_failed setting
                 without re-downloading anything.

    Returns dict with:
        rows, written_rows, unique_urls, failed_rows, output_sheet,
        cache, rows_by_src_row   ← cached for /confirm re-use
    """
    wb = load_workbook(file_path)

    if sheet_src not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_src}' not found. Available: {', '.join(wb.sheetnames)}")

    ws_src = wb[sheet_src]

    if cached_data:
        cache          = cached_data["cache"]
        rows_by_src_row = cached_data["rows_by_src_row"]
    else:
        rows_by_src_row, urls = collect_work(ws_src)
        if not rows_by_src_row:
            raise ValueError("No processable rows found. Ensure Col A (Master ID) and Col B (Stack Group) are filled.")
        cache = fetch_all_filenames(urls, referer, timeout, workers, progress_callback)

    good_rows, failed_rows, failed_src_cells = build_row_results(rows_by_src_row, cache)

    rows_to_write = good_rows if skip_failed else good_rows  # always skip failed rows (row-level rule)
    # Note: we ALWAYS skip failed rows in the output to preserve stack order.
    # The "cancel" path still downloads so user can see the red highlights.

    highlight_source_failures(ws_src, failed_src_cells)
    write_output_sheet(wb, sheet_out, rows_to_write, cache)
    wb.save(output_path)

    total_url_count = sum(len(cells) for cells in rows_by_src_row.values())

    return {
        "rows":          total_url_count,
        "written_rows":  len(rows_to_write),
        "unique_urls":   len(cache),
        "failed_rows":   failed_rows,
        "output_sheet":  sheet_out,
        # cache for /confirm re-use
        "cache":         cache,
        "rows_by_src_row": rows_by_src_row,
    }
